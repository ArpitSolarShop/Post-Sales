"""
Excel Import Script — Imports Arpit Solar Master Sheet into PostgreSQL
Run: cd apps/ai-service && source venv/bin/activate && python import_excel.py
"""

import openpyxl
import psycopg2
import uuid
from datetime import datetime

DB_URL = "postgresql://postgres:password@localhost:5433/master_app"
EXCEL_PATH = "/Users/ratneshmishra/Downloads/_Arpit Solar Master Sheet.xlsx"

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    try:
        return datetime.strptime(str(val).strip(), "%d/%m/%Y").isoformat()
    except:
        return None

def parse_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except:
        return None

def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def map_stage(row_data):
    """Determine ProjectStage from the row's column values"""
    subsidy = clean(row_data.get('subsidyRedeem'))
    pcr = clean(row_data.get('pcr'))
    inst_detail = clean(row_data.get('instDetailSub'))
    dcr = clean(row_data.get('dcr'))
    sealing = clean(row_data.get('sealingIndent'))
    meter = clean(row_data.get('meterTypeSl'))
    doc_status = clean(row_data.get('documentStatus'))
    doc_sub = clean(row_data.get('docSubmitted'))
    plant = clean(row_data.get('plantStatus'))
    inc = clean(row_data.get('incStage'))
    invoice = row_data.get('invoiceDate')
    po = clean(row_data.get('poSigned'))
    survey = clean(row_data.get('surveyStatus'))

    if subsidy and subsidy.lower() not in ['', 'no', 'pending']:
        return 'SUBSIDY_REDEEMED'
    if pcr and pcr.lower() not in ['', 'no', 'pending']:
        return 'PCR_FILED'
    if inst_detail and inst_detail.lower() not in ['', 'no', 'pending']:
        return 'INST_DETAIL_SUBMITTED'
    if dcr and dcr.lower() not in ['', 'no', 'pending']:
        return 'DCR_FILED'
    if sealing and sealing.lower() not in ['', 'no', 'pending']:
        return 'METER_SEALING'
    if doc_status and doc_status.lower() not in ['', 'no', 'pending']:
        return 'DOC_VERIFIED'
    if doc_sub and doc_sub.lower() not in ['', 'no', 'pending']:
        return 'DOC_SUBMITTED'
    if plant and plant.lower() not in ['', 'no', 'pending']:
        return 'PLANT_INSTALLED'
    if inc and inc.lower() not in ['', 'no', 'pending']:
        return 'INC_IN_PROGRESS'
    if invoice:
        return 'INVOICED'
    if po and po.lower() in ['signed', 'yes', 'done']:
        return 'PO_SIGNED'
    if survey and survey.lower() in ['done', 'completed', 'yes']:
        return 'SURVEY_COMPLETED'
    if survey:
        return 'SURVEY_SCHEDULED'
    return 'LEAD_CAPTURED'

def import_mastersheet(conn):
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['MasterSheet']

    # Row 2 is headers, Row 3 is a note, data starts at row 4
    headers = [str(c.value).strip() if c.value else '' for c in ws[2]]
    print(f"Headers: {headers}")

    customers_imported = 0
    projects_imported = 0
    cur = conn.cursor()

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        name = clean(row[0])  # NAME
        if not name:
            continue

        cust_id = str(uuid.uuid4())
        calling_no = clean(row[1])
        mobile = clean(row[2])
        ca_number = clean(row[3])
        division = clean(row[4])
        location = clean(row[9])

        # Insert customer
        cur.execute("""
            INSERT INTO "Customer" (id, name, "callingNo", mobile, "caNumber", division, location, "createdAt", "updatedAt")
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
        """, (cust_id, name, calling_no, mobile, ca_number, division, location))
        customers_imported += 1

        # Build project data
        proj_data = {
            'surveyStatus': clean(row[10]),
            'poSigned': clean(row[13]),
            'invoiceDate': parse_date(row[15]),
            'incStage': clean(row[16]),
            'plantStatus': clean(row[17]),
            'docSubmitted': clean(row[19]),
            'documentStatus': clean(row[20]),
            'meterTypeSl': clean(row[21]),
            'status': clean(row[22]),
            'sealingIndent': clean(row[23]),
            'dcr': clean(row[24]),
            'instDetailSub': clean(row[25]),
            'pcr': clean(row[26]),
            'subsidyRedeem': clean(row[27]),
        }

        stage = map_stage(proj_data)
        proj_id = str(uuid.uuid4())

        cur.execute("""
            INSERT INTO "Project" (
                id, "customerId", capacity, "sourceOfLead", "brandModel", referral,
                "surveyStatus", amount, balance, "poSigned", "invoiceDate",
                "incStage", "plantStatus", "docSubmitted", "documentStatus",
                "meterTypeSl", status, "sealingIndent", dcr, "instDetailSub",
                pcr, "subsidyRedeem", stage, "createdAt", "updatedAt"
            ) VALUES (
                %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, NOW(), NOW()
            )
        """, (
            proj_id, cust_id,
            parse_float(row[5]),   # capacity
            clean(row[6]),         # sourceOfLead
            clean(row[7]),         # brandModel
            clean(row[8]),         # referral
            proj_data['surveyStatus'],
            parse_float(row[11]),  # amount
            parse_float(row[12]),  # balance
            proj_data['poSigned'],
            parse_date(row[15]),   # invoiceDate
            proj_data['incStage'],
            proj_data['plantStatus'],
            proj_data['docSubmitted'],
            proj_data['documentStatus'],
            proj_data['meterTypeSl'],
            proj_data['status'],
            proj_data['sealingIndent'],
            proj_data['dcr'],
            proj_data['instDetailSub'],
            proj_data['pcr'],
            proj_data['subsidyRedeem'],
            stage,
        ))
        projects_imported += 1

    conn.commit()
    print(f"✅ MasterSheet: Imported {customers_imported} customers, {projects_imported} projects")

def import_discom_issues(conn):
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Discom  Issues']
    cur = conn.cursor()
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val = parse_date(row[0])
        name = clean(row[1])
        if not name:
            continue

        cur.execute("""
            INSERT INTO "DiscomIssue" (id, date, "customerName", "caNumber", division, mobile, "issueDesc", remark, status, "createdAt", "updatedAt")
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
        """, (
            str(uuid.uuid4()),
            date_val,
            name,
            clean(row[2]),
            clean(row[3]),
            clean(row[4]),
            clean(row[5]),
            clean(row[6]),
            clean(row[7]) or 'OPEN',
        ))
        count += 1

    conn.commit()
    print(f"✅ Discom Issues: Imported {count} issues")

def import_employees(conn):
    """Create default employees from the Excel schedule sheet"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Sheet11']
    cur = conn.cursor()
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean(row[2]) if len(row) > 2 else None
        if not name:
            continue

        emp_id = str(uuid.uuid4())
        cur.execute("""
            INSERT INTO "Employee" (id, name, role, "isActive", "createdAt", "updatedAt")
            VALUES (%s, %s, 'TECHNICIAN', true, NOW(), NOW())
        """, (emp_id, name))

        mon = clean(row[3]) if len(row) > 3 else None
        tue = clean(row[4]) if len(row) > 4 else None
        wed = clean(row[5]) if len(row) > 5 else None
        thu = clean(row[6]) if len(row) > 6 else None
        fri = clean(row[7]) if len(row) > 7 else None
        sat = clean(row[8]) if len(row) > 8 else None
        sun = clean(row[9]) if len(row) > 9 else None

        cur.execute("""
            INSERT INTO "EmployeeSchedule" (id, "employeeId", monday, tuesday, wednesday, thursday, friday, saturday, sunday)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (str(uuid.uuid4()), emp_id, mon, tue, wed, thu, fri, sat, sun))
        count += 1

    conn.commit()
    print(f"✅ Employees: Imported {count} employees with schedules")


if __name__ == '__main__':
    print("🚀 Starting Arpit Solar Master Sheet Import...")
    conn = psycopg2.connect(DB_URL)

    try:
        # Clear existing data
        cur = conn.cursor()
        cur.execute('DELETE FROM "ActivityLog"')
        cur.execute('DELETE FROM "TaskAssignment"')
        cur.execute('DELETE FROM "Notification"')
        cur.execute('DELETE FROM "NetMeterFile"')
        cur.execute('DELETE FROM "Project"')
        cur.execute('DELETE FROM "DiscomIssue"')
        cur.execute('DELETE FROM "EmployeeSchedule"')
        cur.execute('DELETE FROM "Employee"')
        cur.execute('DELETE FROM "Customer"')
        conn.commit()
        print("🧹 Cleared existing data")

        import_employees(conn)
        import_mastersheet(conn)
        import_discom_issues(conn)

        print("\n🎉 Import complete!")
    except Exception as e:
        conn.rollback()
        print(f"❌ Error: {e}")
        raise
    finally:
        conn.close()
